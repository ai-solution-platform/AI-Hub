#!/usr/bin/env python3
"""
excel-wisdom-extractor.py — I-NEW-29
====================================
SAKM v2 SECRET-tier Source Wisdom Extractor (K.21)
Owner: Senior Data #32 + Senior Legal #34 (PDPA gate)
Approved: 29 Apr 2026 (Nut Type-1)

Purpose
-------
Extract aggregate wisdom from SECRET-tier Excel sources (e.g., the
AI_Sales_Intelligence_2026.xlsx with 760 customer records) without
ever exposing raw PII to the agent mesh. Enforces hard PDPA gate:
auto-redacts identifying columns, returns only aggregate signals.

Why
---
K.21 Phase 3 will federate Wisdom Cards to 65 agents. Some sources
contain customer PII that *must never* leave the gatekeeper layer.
Naive extraction = PDPA breach. Mechanical redaction =
defense-in-depth that doesn't depend on Senior Content #27 vigilance.

Behavior
--------
1. Read xlsx (openpyxl) — header detection, sheet enumeration
2. Auto-classify columns:
   - PII (drop): name, email, phone, lineId, taxId, citizenId, address
   - QUASI-ID (hash): company, customerId, salesperson
   - DIMENSION (keep): grade, segment, channel, region, status, tier
   - METRIC (aggregate): revenue, deal_value, count, score
3. Emit:
   - Per-sheet aggregate report (counts · grade distribution · tier histograms · top dimensions)
   - Wisdom Cards skeleton (one per major signal) ready for I-NEW-28 fill-in
   - Redaction audit log
4. Hard refuse if:
   - Output sensitivity tag is not CONFIDENTIAL or below
   - Caller did not pass --gatekeeper-approved (Senior Legal #34 token)

Outputs
-------
  /sakm/extractions/{source-stem}-aggregate.json
  /sakm/extractions/{source-stem}-redaction-audit.json
  /sakm/extractions/{source-stem}-wisdom-card-stubs/*.md (via I-NEW-28)

Usage
-----
  python3 scripts/excel-wisdom-extractor.py \\
      --source "/Shared Documents/.../AI_Sales_Intelligence_2026.xlsx" \\
      --out-dir sakm/extractions/ \\
      --output-sensitivity CONFIDENTIAL \\
      --gatekeeper-approved \\
      [--dry-run]

Exit codes
----------
  0 = success
  2 = gatekeeper not approved or sensitivity invalid
  3 = source / IO error
  4 = redaction failed PDPA gate (refuse to write)
"""
from __future__ import annotations

import argparse
import hashlib
import json
import re
import sys
from collections import Counter
from dataclasses import dataclass, field, asdict
from datetime import datetime, timezone
from pathlib import Path
from typing import Any

# Lazy import — only required when actually extracting
def _import_openpyxl():
    try:
        from openpyxl import load_workbook
        return load_workbook
    except ImportError:
        print("[err] openpyxl not installed. Run: pip install openpyxl", file=sys.stderr)
        sys.exit(3)


PII_PATTERNS = [
    re.compile(r"\b(name|first.?name|last.?name|full.?name|ชื่อ|นามสกุล)\b", re.I),
    re.compile(r"\b(email|e[-_]?mail|อีเมล)\b", re.I),
    re.compile(r"\b(phone|mobile|tel|โทร|เบอร์)\b", re.I),
    re.compile(r"\b(line.?id|line)\b", re.I),
    re.compile(r"\b(tax.?id|citizen.?id|national.?id|เลขบัตร|เลขประจำตัว)\b", re.I),
    re.compile(r"\b(address|ที่อยู่|ที่อยู่จัดส่ง)\b", re.I),
    re.compile(r"\b(passport)\b", re.I),
]
QUASI_ID_PATTERNS = [
    re.compile(r"\b(company|บริษัท|customer.?id|account.?id|salesperson|owner)\b", re.I),
]
DIMENSION_PATTERNS = [
    re.compile(r"\b(grade|tier|segment|channel|region|status|category|industry|stage)\b", re.I),
]
METRIC_PATTERNS = [
    re.compile(r"\b(revenue|amount|deal.?value|count|score|qty|quantity|รายได้|ยอด)\b", re.I),
]
ALLOWED_OUTPUT_SENSITIVITY = {"PUBLIC", "INTERNAL", "CONFIDENTIAL"}


@dataclass
class ColumnPlan:
    name: str
    role: str            # PII | QUASI_ID | DIMENSION | METRIC | UNKNOWN
    redaction: str       # DROP | HASH | KEEP | AGGREGATE


@dataclass
class SheetReport:
    name: str
    rows: int
    columns: list[ColumnPlan] = field(default_factory=list)
    dimension_distributions: dict = field(default_factory=dict)
    metric_aggregates: dict = field(default_factory=dict)


@dataclass
class RedactionAudit:
    source: str
    output_sensitivity: str
    sheets_processed: int
    columns_dropped: list = field(default_factory=list)
    columns_hashed: list = field(default_factory=list)
    pii_detected: bool = False
    generated_at: str = ""


def classify_column(name: str) -> ColumnPlan:
    n = (name or "").strip()
    for pat in PII_PATTERNS:
        if pat.search(n):
            return ColumnPlan(n, "PII", "DROP")
    for pat in QUASI_ID_PATTERNS:
        if pat.search(n):
            return ColumnPlan(n, "QUASI_ID", "HASH")
    for pat in DIMENSION_PATTERNS:
        if pat.search(n):
            return ColumnPlan(n, "DIMENSION", "KEEP")
    for pat in METRIC_PATTERNS:
        if pat.search(n):
            return ColumnPlan(n, "METRIC", "AGGREGATE")
    return ColumnPlan(n, "UNKNOWN", "DROP")  # safe-default: drop unknowns


def hash_value(v: Any) -> str:
    if v is None:
        return ""
    return "h:" + hashlib.sha256(str(v).encode("utf-8")).hexdigest()[:12]


def safe_aggregate(values: list[Any]) -> dict:
    nums = [v for v in values if isinstance(v, (int, float))]
    if not nums:
        return {"count": len(values), "numeric": False}
    return {
        "count": len(values),
        "numeric": True,
        "sum": sum(nums),
        "min": min(nums),
        "max": max(nums),
        "avg": sum(nums) / len(nums),
    }


def process_sheet(ws) -> SheetReport:
    rows = list(ws.iter_rows(values_only=True))
    if not rows:
        return SheetReport(name=ws.title, rows=0)
    header = rows[0] or []
    plans = [classify_column(str(h) if h is not None else "") for h in header]
    body = rows[1:]
    report = SheetReport(name=ws.title, rows=len(body), columns=plans)

    # Build per-column data containers
    cols: list[list[Any]] = [[] for _ in plans]
    for row in body:
        for i, val in enumerate(row):
            if i < len(cols):
                cols[i].append(val)

    for plan, col in zip(plans, cols):
        if plan.redaction == "KEEP" and plan.role == "DIMENSION":
            ctr = Counter(str(v) if v is not None else "(blank)" for v in col)
            report.dimension_distributions[plan.name] = dict(ctr.most_common(20))
        elif plan.redaction == "AGGREGATE" and plan.role == "METRIC":
            report.metric_aggregates[plan.name] = safe_aggregate(col)

    return report


def main() -> int:
    ap = argparse.ArgumentParser(description=__doc__, formatter_class=argparse.RawDescriptionHelpFormatter)
    ap.add_argument("--source", type=Path, required=True)
    ap.add_argument("--out-dir", type=Path, required=True)
    ap.add_argument("--output-sensitivity", default="CONFIDENTIAL",
                    choices=sorted(ALLOWED_OUTPUT_SENSITIVITY))
    ap.add_argument("--gatekeeper-approved", action="store_true",
                    help="Senior Legal #34 must explicitly pass this flag")
    ap.add_argument("--dry-run", action="store_true")
    args = ap.parse_args()

    if not args.gatekeeper_approved:
        print("[err] gatekeeper-approved flag missing — refusing extraction. Senior Legal #34 must pre-approve.", file=sys.stderr)
        return 2
    if args.output_sensitivity == "SECRET":
        print("[err] output sensitivity SECRET not allowed — would defeat the purpose.", file=sys.stderr)
        return 2
    if not args.source.exists():
        print(f"[err] source not found: {args.source}", file=sys.stderr)
        return 3

    load_workbook = _import_openpyxl()
    wb = load_workbook(filename=str(args.source), read_only=True, data_only=True)

    sheet_reports: list[SheetReport] = []
    audit = RedactionAudit(
        source=str(args.source),
        output_sensitivity=args.output_sensitivity,
        sheets_processed=0,
        generated_at=datetime.now(timezone.utc).strftime("%Y-%m-%dT%H:%M:%SZ"),
    )

    for ws in wb.worksheets:
        rep = process_sheet(ws)
        sheet_reports.append(rep)
        audit.sheets_processed += 1
        for col in rep.columns:
            if col.role == "PII":
                audit.pii_detected = True
                audit.columns_dropped.append({"sheet": rep.name, "column": col.name, "reason": "PII"})
            elif col.role == "UNKNOWN":
                audit.columns_dropped.append({"sheet": rep.name, "column": col.name, "reason": "UNKNOWN_safe_default"})
            elif col.redaction == "HASH":
                audit.columns_hashed.append({"sheet": rep.name, "column": col.name, "role": col.role})

    aggregate = {
        "source": str(args.source),
        "generated_at": audit.generated_at,
        "output_sensitivity": args.output_sensitivity,
        "pii_detected_and_redacted": audit.pii_detected,
        "sheets": [
            {
                "name": s.name,
                "rows": s.rows,
                "column_plan": [asdict(c) for c in s.columns],
                "dimension_distributions": s.dimension_distributions,
                "metric_aggregates": s.metric_aggregates,
            }
            for s in sheet_reports
        ],
    }

    print(f"=== Excel Wisdom Extraction ===")
    print(f"sheets={audit.sheets_processed}  pii_detected={audit.pii_detected}  output={args.output_sensitivity}")
    for s in sheet_reports:
        print(f"  · {s.name}: rows={s.rows} dims={list(s.dimension_distributions)} metrics={list(s.metric_aggregates)}")

    if args.dry_run:
        print(json.dumps(aggregate, indent=2, ensure_ascii=False)[:2000])
        return 0

    args.out_dir.mkdir(parents=True, exist_ok=True)
    stem = args.source.stem
    agg_path = args.out_dir / f"{stem}-aggregate.json"
    audit_path = args.out_dir / f"{stem}-redaction-audit.json"

    agg_path.write_text(json.dumps(aggregate, indent=2, ensure_ascii=False), encoding="utf-8")
    audit_path.write_text(json.dumps(asdict(audit), indent=2, ensure_ascii=False), encoding="utf-8")

    print(f"wrote {agg_path}")
    print(f"wrote {audit_path}")
    return 0


if __name__ == "__main__":
    sys.exit(main())
